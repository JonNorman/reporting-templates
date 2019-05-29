# coding: utf-8
from builtins import dict, input

from cleaners import *
from exhelp import *
from functools import reduce
from validators import *

from copy import copy
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from functools import partial
import logging
import numpy
import openpyxl
import os
import pandas
import re
import shutil
import sys
import warnings
import xlrd

logging.basicConfig(level = logging.INFO, format = '%(message)s')
pandas.options.mode.chained_assignment = None  # default='warn'

# a helper class to provide additional metadata to a Workbook
class WrappedWorkbook:

    def __init__(self, workbook, path = None, start_marker = None, end_marker = None):
        self.workbook = workbook
        self.path = path
        self.start_marker = start_marker
        self.end_marker = end_marker

# a helper class to handle the parsing and processing of subtotal rows
class SubTotal:

    EXTRACTION_PATTERN = r'^<subtotal_(?P<group_size>\d+)_(?P<group_type>[\w\s]+)>$'
    FORMAT = '<subtotal_{group_size}_{group_type}>'

    def __init__(self, cell, group_size, group_type):
        self.cell = cell
        self.group_size = int(group_size)
        self.group_type = group_type

    # return all the cells in this group (not including the subtotal cell)
    def cells_above_in_group(self):
        cells_above = [ self.cell.offset(row = -i) for i in range(self.group_size, 0, -1) ]
        return sorted(cells_above, key = lambda c: c.coordinate)

    @staticmethod
    def from_cell(cell):
        matches = re.match(SubTotal.EXTRACTION_PATTERN, str(cell.value)).groupdict()
        return SubTotal(cell, matches['group_size'], matches['group_type'])

    @staticmethod
    def build_label(group_size, group_type):
        return SubTotal.FORMAT.format(group_size = group_size, group_type = group_type)

# a helper class to handle the parsing and processing of total rows
class Total:
    EXTRACTION_PATTERN = r'^<total_(?P<group_type>[\w\s]+)>$'
    FORMAT = '<total_{group_type}>'

    def __init__(self, cell, group_type):
        self.cell = cell
        self.group_type = group_type

    @staticmethod
    def from_cell(cell):
        matches = re.match(Total.EXTRACTION_PATTERN, str(cell.value)).groupdict()
        return Total(cell, matches['group_type'])

    @staticmethod
    def build_label(group_type):
        return Total.FORMAT.format(group_type = group_type)


def get_input_paths(inputs_dir):
    # read input file(s) from inputs directory
    for filename in os.listdir(inputs_dir):
        file_path = os.path.join(inputs_dir, filename)
        if os.path.isfile(file_path) \
            and filename[:2] != '~$' \
            and filename[0] != '.':
            yield file_path

def initialise_workbook(template_path, output_path):
    logging.debug('Copying template from "{0}" to "{1}"...'.format(template_path, output_path))

    if (os.path.isfile(output_path)):
        print('File "{0}" already exists!'.format(output_path))

        if query_yes_no('Do you want to proceed and overwrite this file?'):
            logging.info('Overwriting "{0}"...'.format(output_path))
        else:
            logging.info('Not overwriting file; ignoring this report...')
            return None
    else:
        logging.debug('Creating report file  "{0}"...'.format(output_path))

    shutil.copyfile(template_path, output_path)

    # there is an annoying feature-missing log that gets output
    logging.captureWarnings(True)
    logging.disable(logging.ERROR)
    workbook = openpyxl.load_workbook(filename = output_path)
    logging.disable(logging.NOTSET)
    return WrappedWorkbook(workbook = workbook, path = output_path)

def extract_line_item_id(line_item_name):
    line_item_id_pattern = r'.*?(?P<line_item_id>ORD-\d+-\d+-\d+).*'
    matches = re.match(line_item_id_pattern, line_item_name)
    if matches:
        return matches.group('line_item_id')

def extract_order_id(line_item_name):
    order_id_pattern = r'.*?(?P<order_id>ORD-\d+).*'
    matches = re.match(order_id_pattern, line_item_name)
    if matches:
        return matches.group('order_id')

# Line items are grouped if they have the same start/end date and line-item-id
def group_line_items(df):

    def get_group_id(index):

        row = df.loc[index]

        line_item_id = extract_line_item_id(row['Line Item'])
        order_id = extract_order_id(row['Line Item'])
        extracted_id = line_item_id if line_item_id else order_id
        return '{}-{}-{}'.format(str(row['Line item start date']),
                                 str(row['Line item end date']),
                                 extracted_id)

    return (df.sort_values(by = ['Line item start date', 'Line Item', 'Creative Size'], inplace = False)
              .groupby(by = get_group_id, sort = True))

def write_data(df, wrapped_workbook):
    worksheet = wrapped_workbook.workbook.active

    # get the start and end tags
    first_row_start = get_cell_by_regex(worksheet, '<data_start>')
    first_row_end = get_cell_by_regex(worksheet, '<data_end>')
    header_start = get_cell_by_regex(worksheet, '<header_start>')

    if not all([first_row_start, first_row_end, header_start]):
        return offer_clean_exit(wrapped_workbook.path)

    # write the headers
    headers = list(df.columns.values)
    for column, header in enumerate(headers):
        header_start.offset(column = column).value = header

    total_row = [ Total.build_label(h) for h in headers ]
    blank_row = [None] * len(headers)

    def replace_nan(row, replacement = 'n/a'):
        return ['n/a' if type(cell) is float and numpy.isnan(cell) else cell for cell in row ]

    rows_to_write = []
    for _, line_items in group_line_items(df):

        group_size = len(line_items)

        # provide a new line above groups (if they aren't the first row and there isn't already one)
        if rows_to_write and group_size > 1 and rows_to_write[-1] != blank_row:
            rows_to_write.append(blank_row)

        # write each row of each group
        rows = dataframe_to_rows(line_items, index = False, header = False)
        for row in rows:
            rows_to_write.append(replace_nan(row))

        # if we have more than one line in a group, write a subtotal row
        if group_size > 1:
            subtotal_row = [ SubTotal.build_label(group_size, h) for h in headers ]
            rows_to_write.append(subtotal_row)
            rows_to_write.append(blank_row)

    # add a blank row after all the data has been written, then add a total row
    rows_to_write.append(blank_row)
    rows_to_write.append(total_row)

    # add a row for the total border
    rows_to_write.append([ '<total_bar>' for h in headers] )

    # now write them to the worksheet
    for row, values in enumerate(rows_to_write):
        for column, value in enumerate(values):
            first_row_start.offset(row = row, column = column).value = value

    return WrappedWorkbook(workbook = wrapped_workbook.workbook,
                           path = wrapped_workbook.path,
                           start_marker = first_row_start,
                           end_marker = first_row_end.offset(row = len(rows_to_write) - 1))

def apply_styling(wrapped_workbook, image_path, columns_to_merge = []):

    logging.debug('Applying styling...')

    worksheet = wrapped_workbook.workbook.active
    data_start, data_end = wrapped_workbook.start_marker, wrapped_workbook.end_marker

    # copy the styles down each column
    for column in worksheet.iter_cols(min_col = data_start.col_idx,
                                      max_col = data_end.col_idx,
                                      min_row = data_start.row,
                                      max_row = data_end.row):

        # use the first cell in each column as a format for the rest
        template = column[0]
        for cell in column:
            cell.font = copy(template.font)
            cell.alignment = copy(template.alignment)
            if template.is_date:
                cell.number_format = 'dd/mm/yyyy'
            else:
                cell.number_format = copy(template.number_format)

    # embolden all of the tags
    for tag in get_cells_by_regex(worksheet, r'^<[\w\s]+>$'):
        tag.font = update_font(tag.font, {'bold': True})

        # specifically format the total bar
        if tag.value == r'<total_bar>':
            tag.value = None
            tag.fill = PatternFill(patternType = 'solid',
                                   fgColor = 'F2F2F2')

    # add the picture and limit the size
    img = Image(image_path, coordinates = ((0,0), (1,1)), size = (80, 80))
    worksheet.add_image(img, get_cell_by_regex(worksheet, r'<icon>').coordinate)

    # merge the rows within groups for the columns specified
    all_subtotals = [ SubTotal.from_cell(cell) for cell in get_cells_by_regex(worksheet, SubTotal.EXTRACTION_PATTERN) ]
    subtotals_for_merging = [ subtotal for subtotal in all_subtotals if subtotal.group_type in columns_to_merge ]

    def get_range(from_first, to_last):
        return '{}:{}'.format(from_first.coordinate, to_last.coordinate)
        worksheet.merge_cells(merge_range)

    for subtotal in subtotals_for_merging:
        group = sorted(subtotal.cells_above_in_group(), key = lambda c: c.coordinate)
        worksheet.merge_cells(get_range(group[0], group[-1]))

    # merge the order_id cells
    order_id_start = get_cell_by_regex(worksheet, r'<order_id>')
    order_id_end = worksheet.cell(row = data_end.row, column = order_id_start.col_idx)
    order_id_range = get_range(order_id_start, order_id_end)
    worksheet.merge_cells(order_id_range)

    # add a border to the order_id block
    for row in worksheet[order_id_range]:
        for cell in row:
            cell.border = get_border(top_left = order_id_start,
                                     bottom_right = order_id_end,
                                     cell = cell)
            cell.alignment = Alignment(horizontal = 'center',
                                       vertical = 'center',
                                       textRotation = 90)

    # add a border to the data block
    for row in worksheet['{}:{}'.format(data_start.coordinate, data_end.coordinate)]:
        for cell in row:
            cell.border = get_border(top_left = data_start,
                                     bottom_right = data_end,
                                     cell = cell)

    return wrapped_workbook

def write_totals(wrapped_workbook):

    # open up the workbook
    worksheet = wrapped_workbook.workbook.active
    data_start = wrapped_workbook.start_marker
    data_end = wrapped_workbook.end_marker

    # replace the totals first; subtotalled rows need not be counted twice
    total_cells = [ c for c in get_cells_by_regex(worksheet, Total.EXTRACTION_PATTERN) ]
    subtotal_cells = [ c for c in get_cells_by_regex(worksheet, SubTotal.EXTRACTION_PATTERN) ]

    def identify_non_subtotalled_cells(candidates, subtotal_cells):

        # determine which cells are covered by subtotals (including the subtotals themselves)
        subtotalled_cells = []
        for subtotal_cell in subtotal_cells:
            subtotal = SubTotal.from_cell(subtotal_cell)
            for i in range(0, subtotal.group_size + 1, 1):
                subtotalled_cells.append(subtotal_cell.offset(row = -i))

        # any cells not caught (and non blank) need to be captured
        return [ c for c in candidates if c not in (subtotalled_cells + subtotal_cells) and c.value ]

    for cell in total_cells:
        total = Total.from_cell(cell)
        if total.group_type == 'Ad server CTR':
            total.cell.value = '={}/{}'.format(total.cell.offset(column = -1).coordinate,
                                              total.cell.offset(column = -2).coordinate)
        elif total.group_type == 'Delivery Indicator':
            total.cell.value = 'TOTAL'

        elif total.group_type in ['Ad server clicks', 'Ad server impressions']:
            column_above = [ c for c in cells_between(data_start, total.cell.offset(row = - 1)) if c.column == total.cell.column ]
            subtotals_above = [ s for s in subtotal_cells if s in column_above ]

            non_subtotalled_cells = identify_non_subtotalled_cells(column_above, subtotals_above)
            total.cell.value = get_add_formula(subtotals_above + non_subtotalled_cells)

    for cell in subtotal_cells:
        subtotal = SubTotal.from_cell(cell)

        if subtotal.group_type == 'Ad server CTR':
            subtotal.cell.value = '={}/{}'.format(subtotal.cell.offset(column = -1).coordinate,
                                              subtotal.cell.offset(column = -2).coordinate)
        elif subtotal.group_type == 'Delivery Indicator':
            subtotal.cell.value = 'Total'

        elif subtotal.group_type in ['Ad server clicks', 'Ad server impressions']:
            group = subtotal.cells_above_in_group()
            subtotal.cell.value = get_sum_formula(group[0], group[-1])

    return wrapped_workbook

def replace_order_id(wrapped_workbook, order_id):
    worksheet = wrapped_workbook.workbook.active
    order_id_cell = get_cell_by_regex(worksheet, r'<order_id>')
    order_id_cell.value = order_id

    return wrapped_workbook

def remove_extra_tags(wrapped_workbook):
    for tag in get_cells_by_regex(wrapped_workbook.workbook.active, r'^<[\w\s]+>$'):
        tag.value = None

    return wrapped_workbook

def save_workbook(wrapped_workbook):
    wrapped_workbook.workbook.save(wrapped_workbook.path)
    return wrapped_workbook

def main():

    report_columns = ['Line Item',
                      'Creative',
                      'Delivery Indicator',
                      'Line item start date',
                      'Line item end date',
                      'Goal quantity',
                      'Creative Size',
                      'DAP Native Format',
                      'Ad server impressions',
                      'Ad server clicks',
                      'Ad server CTR'
                      ]

    report_columns_trimmed_and_ordered = ['Line Item',
                                          'Line item start date',
                                          'Line item end date',
                                          'Goal quantity',
                                          'Creative Size',
                                          'Delivery Indicator',
                                          'Ad server impressions',
                                          'Ad server clicks',
                                          'Ad server CTR']

    columns_to_merge = ['Delivery Indicator',
                        'Line item start date',
                        'Line item end date',
                        'Goal quantity']

    inputs = {
        'Sheet Name'              : 'Report data',
        'Inputs Directory'        : 'inputs',
        'Outputs Directory'       : 'outputs',
        'Template'                : os.path.join('assets','template.xlsx'),
        'Icon'                    : os.path.join('assets','icon.png'),
        'Goal Quantity Threshold' : 1000,
        'Required Columns'        : ', '.join(report_columns)
    }

    validators = [partial(validate_column_names, required_columns = report_columns)]
    cleaners = [partial(drop_row_with_value_in_column, column_name = 'Line Item',
                                                       value = 'Total',
                                                       exact_match = True),
                partial(drop_row_with_value_in_column, column_name ='Line Item',
                                                       value = 'TEST',
                                                       exact_match = False),
                partial(replace_value_below_threshold_with_nan, column_name = 'Goal quantity',
                                                                threshold = inputs['Goal Quantity Threshold']),
                partial(replace_column_value, column_name = 'Creative Size',
                                              pattern = '1 x 1',
                                              replacement = 'pageskin'),
                partial(replace_column_extract, column_name = 'Creative Size',
                                              extract_value_from_row = lambda row: row['DAP Native Format'] if row['Creative Size'] == 'Native' and row['DAP Native Format'] != '-' else row['Creative Size']),
              partial(drop_columns_not_required, required_columns = report_columns_trimmed_and_ordered),
                partial(replace_datetime_with_date, column_name = 'Line item start date'),
                partial(replace_datetime_with_date, column_name = 'Line item end date'),
                partial(reorder_columns, ordered_columns = report_columns_trimmed_and_ordered)]

    log_divider(symbol = '*')
    logging.info('Running the report formatter with the following inputs:')
    for prop, value in inputs.items(): logging.info('\t{:25s}: {}'.format(prop, value))
    log_divider(symbol = '*')

    summaries = []
    for input_path in get_input_paths(inputs['Inputs Directory']):

        logging.info('Found input "{}" '.format(input_path))
        if not query_yes_no('Want to format this report?'):
            logging.info('Okay, looking for other inputs...')
            continue

        df_raw = pandas.read_excel(io = input_path, sheetname = inputs['Sheet Name'])

        # if the input is invalid then continue with the next input_path
        if not all([ validator(df_raw) for validator in validators ]):
            logging.error('Input report is invalid; report will be skipped.'.format(input_path))
            continue

        logging.info('\nCleaning data...')
        df_clean = reduce(lambda df, cleaner: cleaner(df), cleaners, df_raw)

        output_filename = 'formatted_' + os.path.basename(input_path)
        output_path = os.path.join(inputs['Outputs Directory'], output_filename)
        wrapped_workbook = initialise_workbook(template_path = inputs['Template'],
                                               output_path = output_path)

        order_id = extract_order_id(df_raw['Line Item'][0])

        # if the intialisation has been cancelled by the user, go to next file
        if not wrapped_workbook:
            continue

        # each of these functions returns a WrappedWorkbook or None
        transformations = [partial(write_data, df_clean),
                           partial(apply_styling, image_path = inputs['Icon'], columns_to_merge = columns_to_merge),
                           write_totals,
                           partial(replace_order_id, order_id = order_id),
                           remove_extra_tags,
                           save_workbook]

        reduce(lambda wrapped_workbook, f: bind(wrapped_workbook, f), transformations, wrapped_workbook)

        logging.info('Workbook formatting complete!\n')
        log_divider(symbol = '*')

        pandas.set_option('display.float_format', lambda x: '%.3f' % x)
        summary = [('Report Name', output_filename),
                   ('Order ID', order_id),
                   ('Total Ad server impressions', df_clean['Ad server impressions'].sum()),
                   ('Total Ad server clicks', df_clean['Ad server clicks'].sum())]

        summaries.append(summary)

    log_divider()
    logging.info('IMPORTANT: Below is a summary of the input report(s). ' +
                 'Please ensure that these totals match those in the final report(s).')
    logging.info('Note: these totals will not take into account any rows removed ' +
                 'during the cleansing process')
    log_divider()
    for summary in summaries:
        for prop, value in summary:
            logging.info('{:30s}: "{}"'.format(prop, value))
        log_divider()
        logging.info('')

    input('All inputs have been processed! Press ENTER to exit...')
    print('')

if __name__ == '__main__':
    main()
