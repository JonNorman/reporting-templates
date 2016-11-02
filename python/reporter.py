from cleaners import *
from exhelp import *
from validators import *

from copy import copy
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from functools import partial
import logging
import openpyxl
import os
import pandas
import re
import shutil
import sys
import xlrd

logging.basicConfig(level = logging.INFO, format = '> %(message)s')

def read_inputs(inputs_dir = 'inputs',
                input_sheet_name = 'Report data',
                validators = [],
                cleaners = []):

    # read input file(s) from input directory into a raw df
    logging.info('Searching for excel input files...')

    def xl_inputs():
        # for each directory
        for (dirpath, dirnames, filenames) in os.walk(inputs_dir):
            for filename in filenames:
                if filename[:2] != '~$':
                    yield os.path.join(dirpath,filename)

    # for each file in the xl_filenames, validate and clean
    for xl_input in xl_inputs():

        logging.info('Reading in file "{0}"...'.format(xl_input))
        df_raw = pandas.read_excel(io = xl_input, sheetname = input_sheet_name)

        # if validators have been provided, use them
        logging.info('Validating data...')
        if validators:
            if all([ validator(df_raw) for validator in validators ]):
                logging.debug('Input data is valid according to provided validator(s).')
            else:
                logging.error('Input data is invalid according to provided validator;' +
                      'file {0} will be skipped.'.format(xl_input)
                      )
                continue

        # if cleaners have been provided, use them
        logging.info('Cleaning data...')
        if cleaners:
            df_clean = reduce(lambda df, cleaner: cleaner(df), cleaners, df_raw)
        else:
            df_clean = df_raw

        yield (df_clean, os.path.basename(xl_input))

def create_workbook_from_template(template_path, output_path):
    logging.info('Copying template from "{0}" to "{1}"...'.format(template_path, output_path))

    if (os.path.isfile(output_path)):
        overwrite = query_yes_no('* File "{0}" already exists. Confirm overwrite?'.format(output_path))
        if overwrite:
            logging.info('Overwriting "{0}"...'.format(output_path))
        else:
            logging.info('Not overwriting file; ignoring this report...')
            return None
    else:
        logging.info('Creating report file  "{0}"...'.format(output_path))

    shutil.copyfile(template_path, output_path)
    return openpyxl.load_workbook(output_path)

# Line items are grouped if they have the same start/end date and ORD-XXXXXXX-X-X number
def identify_line_item_groups(df):

    def extract_id(line_item_name):
        line_item_id_pattern = '.*?(ORD-\d+-\d+-\d+).*'
        matches = re.match(line_item_id_pattern, line_item_name)
        return matches.group(1)

    df['Line Item Group'] = df['Line Item'].map(extract_id)
    df['Line Item Group'] = df['Line Item Group'] + '-' + df['Line item start date'].map(str)
    df['Line Item Group'] = df['Line Item Group'] + '-' + df['Line item end date'].map(str)
    return df


def offer_clean_exit(output_path):
    logging.error('Skipping report creation for report "{}".'.format(output_path))
    if query_yes_no('* Output file "{}" is being abandoned - want me to delete it?'.format(output_path)):
        logging.info('Deleting file {}...'.format(output_path))
        os.remove(output_path)
    return

def write_data(df, workbook, output_path):
    worksheet = workbook._sheets[0]

    # check the tags
    expected_tags = ['<header_start>', '<header_end>', '<data_start>', '<data_end>', '<order_id>']
    tags = get_tags(worksheet, expected_tags, unique = True)

    if tags is None:
        return offer_clean_exit(output_path)

    # group alike line items together and prepare to write them to the worksheet
    line_item_groups = identify_line_item_groups(df) \
                        .sort_values(by = 'Creative size') \
                        .groupby(by = ['Line Item Group'], sort = ['Line item end date', 'Line item start date'])

    line_item_groups = { group_id: line_items.drop('Line Item Group', 1) for (group_id, line_items) in  line_item_groups }

    # replace the header row with the headers from the groups
    logging.debug('Writing headers to output on line {}'.format(tags['<header_start>'].row))
    headers = list(line_item_groups.itervalues().next().columns.values)
    write_row_between_tags(tags['<header_start>'], tags['<header_end>'], headers)

    # write the data rows
    start_tag, end_tag = tags['<data_start>'], tags['<data_end>']
    _, data_width = distance(start_tag, end_tag)

    # Beware of changing when new lines are written. The distinction between
    # a blank line and a 'total' line at the bottom of a block of lines is
    # important for determining which lines are grouped and which are not.
    for _, line_items in line_item_groups.iteritems():

        is_grouped = len(line_items) > 1

        # add a blank line above each group (unless it is the first group)
        if is_grouped and start_tag != tags['<data_start>']:
            start_tag, end_tag = write_row_between_tags(start_tag, end_tag, [None] * data_width)

        # write each row
        for row in dataframe_to_rows(line_items, index = False, header = False):
            start_tag, end_tag = write_row_between_tags(start_tag, end_tag, row)

        # if we have more than one line item in a group, then we need to provide a subtotal row
        if is_grouped:
            logging.debug('Writing total row on  line {}...'.format(start_tag.row))
            total_tags  = ['<subtotal_label>', '<subtotal_impressions>', '<subtotal_clicks>', '<subtotal_clickthrough>']
            filler = [None] * (data_width - len(total_tags))
            start_tag, end_tag = write_row_between_tags(start_tag, end_tag, filler + total_tags)
            start_tag, end_tag = write_row_between_tags(start_tag, end_tag)


    # add a blank row after all the data has been written
    start_tag, end_tag = write_row_between_tags(start_tag, end_tag)

    # add a grand total row
    total_tags  = ['<total_label>', '<total_impressions>', '<total_clicks>', '<total_clickthrough>']
    filler = [None] * (data_width - len(total_tags))
    start_tag, end_tag = write_row_between_tags(start_tag, end_tag, filler + total_tags)

    return (tags['<data_start>'], end_tag)

def apply_styling(workbook, data_start, data_end, image_path):

    logging.info('Applying styling...')

    worksheet = workbook._sheets[0]
    length, width = distance(data_start, data_end)

    for column in worksheet.iter_cols(min_col = data_start.col_idx,
                                  max_col = data_end.col_idx,
                                  min_row = data_start.row,
                                  max_row = data_end.row):

        # use the first cell in each column as a format for the rest
        template = column[0]

        for cell in column:
            cell.font = copy(template.font)
            cell.alignment = copy(template.alignment)
            cell.number_format = copy(template.number_format)

    # embolden all of the tags
    for tag_name, tags in get_tags(worksheet, unique = False).iteritems():
        for tag in tags:
            tag.font = Font(name=tag.font.name, size=tag.font.size, bold=True,
                            italic=tag.font.italic, vertAlign=tag.font.vertAlign,
                            underline=tag.font.underline, strike=tag.font.strike,
                            color=tag.font.color)

    # add the picture and limit the size
    tags = get_tags(worksheet, unique = True)
    img = Image(image_path, coordinates = ((0,0), (1,1)), size = (80, 80))
    worksheet.add_image(img, tags['<icon>'].coordinate)

    # border around data
    def get_borders(cell, top_left, bottom_right):
        border = Side(style='medium')
        return {
            'left': border if cell.col_idx == top_left.col_idx else Side(None),
            'right': border if cell.col_idx == bottom_right.col_idx else Side(None),
            'top': border if cell.row == top_left.row else Side(None),
            'bottom': border if cell.row == bottom_right.row else Side(None)
        }

    # merge the order_id cell down and add a border
    order_id = tags['<order_id>']
    bottom = worksheet.cell(row = data_end.row, column = order_id.col_idx)

    merge_range = '{}:{}'.format(order_id.coordinate, bottom.coordinate)
    worksheet.merge_cells(merge_range)

    for row in worksheet['{}:{}'.format(order_id.coordinate, bottom.coordinate)]:
        for cell in row:
            cell.border = Border(**get_borders(cell, order_id, bottom))


    for row in worksheet['{}:{}'.format(data_start.coordinate, data_end.coordinate)]:
        for cell in row:
            cell.border = Border(**get_borders(cell, data_start, data_end))

    return (data_start, data_end)

def replace_tags(workbook, data_start, data_end):

    # open up the workbook
    worksheet = workbook._sheets[0]
    logging.info('Replacing markup tags...')

    def sum_subtotal_column(subtotal_tag):
        subtotal_tag.value = '=SUM({}:{})'.format(
            subtotal_tag.offset(row = -1).coordinate,
            first_non_blank_cell_above(subtotal_tag).coordinate
        )

    def sum_total_column(total_tag, subtotal_tags):

        # grouped lines have a total row directly below them
        # ungrouped lines have blank lines
        def is_ungrouped(cell):
            tags = subtotal_tags + [total_tag]

            # we don't care about these values
            if cell.value is None or cell.value in tags:
                return False

            # now burrow down to what ultimately sits below...
            sentry = cell
            while not(sentry.value is None or sentry in tags):
                sentry = sentry.offset(row = 1)

            return sentry.value is None

        tags = subtotal_tags + [total_tag]
        column = [ worksheet.cell(row = row, column = total_tag.col_idx) for row in range(data_start.row, total_tag.row) ]
        ungrouped_cells = [ cell for cell in column if is_ungrouped(cell) ]

        cells_to_sum = ungrouped_cells + subtotal_tags
        coordinates_to_sum = map(lambda cell: cell.coordinate, cells_to_sum)

        # we need to sum subtotals + non-grouped line items
        total_tag.value = '={}'.format('+'.join(coordinates_to_sum))

    expected_total_tags = ['<total_label>',
                           '<total_impressions>',
                           '<total_clicks>',
                           '<total_clickthrough>']
    expected_subtotal_tags = ['<subtotal_label>',
                              '<subtotal_impressions>',
                              '<subtotal_clicks>',
                              '<subtotal_clickthrough>']

    totals = get_tags(worksheet, expected_total_tags, unique = True)
    subtotals = get_tags(worksheet, expected_subtotal_tags, unique = False)

    if totals is None or subtotals is None:
        return offer_clean_exit(output_path)

    # replace totals first
    logging.debug('Writing grand total values to totals on row {}'.format(totals['<total_label>'].row))

    totals['<total_label>'].value = 'TOTAL'
    sum_total_column(totals['<total_impressions>'], subtotals['<subtotal_impressions>'])
    sum_total_column(totals['<total_clicks>'], subtotals['<subtotal_clicks>'])
    totals['<total_clickthrough>'].value = '={}/{}'.format(
        totals['<total_clicks>'].coordinate,
        totals['<total_impressions>'].coordinate)


    # replace clickthrough subtotals first - using total tags on same row
    for clickthrough in subtotals['<subtotal_clickthrough>']:
        logging.debug('Writing clickthrough rate subtotal formulae to {}'.format(clickthrough.coordinate))
        impressions_tag = filter(lambda total: total.row == clickthrough.row, subtotals['<subtotal_impressions>'])[0]
        clicks_tag = filter(lambda total: total.row == clickthrough.row, subtotals['<subtotal_clicks>'])[0]
        clickthrough.value = '={}/{}'.format(clicks_tag.coordinate, impressions_tag.coordinate)

    # now replace each of the subtotals columns
    for tag in subtotals['<subtotal_impressions>']:
        logging.debug('Writing impressions subtotal formulae to {}'.format(tag.coordinate))
        sum_subtotal_column(tag)

    for tag in subtotals['<subtotal_clicks>']:
        logging.debug('Writing clicks subtotal formulae to {}'.format(tag.coordinate))
        sum_subtotal_column(tag)

    # now replace the total labels
    for tag in subtotals['<subtotal_label>']:
        logging.debug('Writing "Total" labels to {}'.format(tag.coordinate))
        tag.value = 'Total'

    # now replace the order_id tag
    order_id = re.compile('.*?(ORD-\d+)-.*').match(data_start.value).group(1)
    logging.info('Determined Order ID as "{}"; writing to output.'.format(order_id))
    tags = get_tags(worksheet)
    worksheet[tags['<order_id>'].coordinate] = order_id

    return workbook

def main():

    required_columns = ['Line Item',
                           'Creative size',
                           'Line item start date',
                           'Line item end date',
                           'Goal quantity',
                           'Delivery indicator',
                           'Ad server impressions',
                           'Ad server clicks',
                           'Ad server CTR']

    validators = [partial(column_names, required_columns = required_columns)]

    cleaners = [partial(drop_row_with_value_in_column, column_name = 'Line Item',
                                                       value = 'Total',
                                                       exact_match = True),
                partial(drop_row_with_value_in_column, column_name ='Line Item',
                                                       value = 'TEST',
                                                       exact_match = False),
                partial(drop_columns_not_required, required_columns = required_columns),
                partial(replace_goal_quantity_with_nan, column_name = 'Goal quantity',
                                                        threshold = 1000)]

    for df, original_filename in read_inputs(validators = validators, cleaners = cleaners):
        output_path = 'outputs/' + original_filename

        workbook = create_workbook_from_template(template_path = 'template.xlsx', output_path = output_path)

        if workbook:
            start_data, end_data = write_data(df, workbook, output_path)
            all([
                apply_styling(workbook, start_data, end_data, 'icon.png'),
                replace_tags(workbook, start_data, end_data),
                workbook.save(output_path)]
                )

if __name__ == '__main__':
    main()
